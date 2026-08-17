#!/usr/bin/env python3
"""
Algeria Wilayas Comprehensive Dataset Compiler
==============================================
Compiles data from 5 parallel scraping agents into a professional Excel spreadsheet.
Covers all 58 Algerian wilayas (administrative provinces) with complete data.
"""

import json
import os
import sys
from datetime import datetime

# Add xlsx skill to path for templates
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
if XLSX_SKILL_DIR not in sys.path:
    sys.path.insert(0, XLSX_SKILL_DIR)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Import base template styles
from templates.base import (
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_body, font_caption, font_subheader,
    fill_header, fill_data_row, fill_total,
    align_title, align_header, align_number, align_text,
    border_header, border_total,
    COLUMN_WIDTHS, ROW_HEIGHTS, PRIMARY, PRIMARY_LIGHT, NEUTRAL_900,
    NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0, HEADER_TEXT,
    create_bar_chart, create_pie_chart, setup_chart_titles, apply_chart_colors, apply_pie_colors,
    CHART_COLORS, use_palette_explicit
)

# Use professional palette
use_palette_explicit("professional")

# ============================================================
# COMPLETE WILAYAS DATA (Compiled from all 5 parallel agents)
# ============================================================

ALL_WILAYAS = [
    # Wilayas 1-12 (Agent 1: Northern/Coastal)
    {
        "code": 1, "name_arabic": "أدرار", "name_french": "Adrar",
        "capital": "Adrar", "area_km2": 42739, "population": 197000,
        "districts": 6, "communes": 16,
        "coordinates": {"lat": 27.88056, "lng": -0.29722},
        "region": "Sahara", "established": "1974",
        "economy": ["Agriculture", "Date palm cultivation", "Phosphate mining", "Oil & gas"],
        "landmarks": ["Touat oasis region", "Gourara region", "Ancient ksars"]
    },
    {
        "code": 2, "name_arabic": "الشلف", "name_french": "Chlef",
        "capital": "Chlef", "area_km2": 4791, "population": 1356151,
        "districts": 13, "communes": 35,
        "coordinates": {"lat": 36.1667, "lng": 1.3333},
        "region": "North (Coastal)", "established": "1968",
        "economy": ["Agriculture", "Glass production", "Hydroelectric power"],
        "landmarks": ["Ténès coastal city", "Dahra Mountains", "Mediterranean coastline"]
    },
    {
        "code": 3, "name_arabic": "الأغواط", "name_french": "Laghouat",
        "capital": "Laghouat", "area_km2": 25057, "population": 479200,
        "districts": 10, "communes": 24,
        "coordinates": {"lat": 33.45, "lng": 2.87},
        "region": "Highlands (Steppe)", "established": "1957",
        "economy": ["Natural gas production", "Solar power", "Livestock"],
        "landmarks": ["Hassi R'Mel gas field", "Saharan Atlas Mountains", "CPV solar station"]
    },
    {
        "code": 4, "name_arabic": "أم البواقي", "name_french": "Oum El Bouaghi",
        "capital": "Oum El Bouaghi", "area_km2": 6094, "population": 558500,
        "districts": 12, "communes": 29,
        "coordinates": {"lat": 35.68, "lng": 7.12},
        "region": "Highlands (Aurès)", "established": "1974",
        "economy": ["Agriculture", "Trade", "Light industry", "Cereal farming"],
        "landmarks": ["Ain M'lila plain", "Aures region border"]
    },
    {
        "code": 5, "name_arabic": "باتنة", "name_french": "Batna",
        "capital": "Batna", "area_km2": 12192, "population": 1128030,
        "districts": 21, "communes": 61,
        "coordinates": {"lat": 35.55, "lng": 6.17},
        "region": "Highlands (Aurès)", "established": "1968",
        "economy": ["Heavy industry", "Chemical industry", "Textile manufacturing"],
        "landmarks": ["Timgad (UNESCO)", "Belezma National Park", "Aurès Mountains"]
    },
    {
        "code": 6, "name_arabic": "بجاية", "name_french": "Béjaïa",
        "capital": "Béjaïa", "area_km2": 3268, "population": 984050,
        "districts": 19, "communes": 52,
        "coordinates": {"lat": 36.75, "lng": 5.067},
        "region": "North (Kabylie)", "established": "1974",
        "economy": ["Port activities", "Commerce", "Tourism", "Fishing"],
        "landmarks": ["Gouraya National Park", "Béjaïa Port", "Pic des Singes"]
    },
    {
        "code": 7, "name_arabic": "بسكرة", "name_french": "Biskra",
        "capital": "Biskra", "area_km2": 9576, "population": 869215,
        "districts": 11, "communes": 27,
        "coordinates": {"lat": 34.867, "lng": 5.75},
        "region": "Sahara Edge (Ziban)", "established": "1957",
        "economy": ["Date palm cultivation", "Agriculture (oases)", "Thermal tourism"],
        "landmarks": ["Tolga date palm groves", "El-Kantara gorge", "Hammat Essalihine springs"]
    },
    {
        "code": 8, "name_arabic": "بشار", "name_french": "Béchar",
        "capital": "Béchar", "area_km2": 50400, "population": 315600,
        "districts": 13, "communes": 31,
        "coordinates": {"lat": 31.63, "lng": -2.23},
        "region": "Southwest (Saoura/Sahara)", "established": "1974",
        "economy": ["Mining (coal, iron ore)", "Oasis farming", "Energy production"],
        "landmarks": ["Saoura region", "Taghit sand dunes", "Beni Abbes oasis"]
    },
    {
        "code": 9, "name_arabic": "البليدة", "name_french": "Blida",
        "capital": "Blida", "area_km2": 1696, "population": 1009000,
        "districts": 10, "communes": 25,
        "coordinates": {"lat": 36.48, "lng": 2.83},
        "region": "North (Mitidja)", "established": "1968",
        "economy": ["Agriculture (citrus)", "Food processing", "Education"],
        "landmarks": ["Chréa National Park", "Tell Atlas Mountains", "Mitidja plain"]
    },
    {
        "code": 10, "name_arabic": "البويرة", "name_french": "Bouira",
        "capital": "Bouira", "area_km2": 4435, "population": 704462,
        "districts": 12, "communes": 45,
        "coordinates": {"lat": 36.38, "lng": 3.90},
        "region": "Highlands (Kabylie border)", "established": "1974",
        "economy": ["Agriculture", "Light industry", "Transport hub", "Winter sports"],
        "landmarks": ["Djurdjura National Park", "Djurdjura Mountains", "Ski resorts"]
    },
    {
        "code": 11, "name_arabic": "تمنراست", "name_french": "Tamanrasset",
        "capital": "Tamanrasset", "area_km2": 228419, "population": 191800,
        "districts": 6, "communes": 23,
        "coordinates": {"lat": 22.78, "lng": 5.52},
        "region": "Deep Sahara (Hoggar)", "established": "1957",
        "economy": ["Desert tourism", "Trans-Saharan trade", "Crafts", "Pastoralism"],
        "landmarks": ["Hoggar Mountains", "Mount Tahat (2918m)", "Assekrem plateau"]
    },
    {
        "code": 12, "name_arabic": "تبسة", "name_french": "Tébessa",
        "capital": "Tébessa", "area_km2": 13878, "population": 697100,
        "districts": 12, "communes": 28,
        "coordinates": {"lat": 35.41, "lng": 8.08},
        "region": "Southeast (Highlands/Tunisian border)", "established": "1968",
        "economy": ["Phosphate mining", "Agriculture", "Trade with Tunisia"],
        "landmarks": ["Tembisa Roman ruins", "Byzantine fortress", "Archaeological museum"]
    },
    
    # Wilayas 13-26 (Agent 2: Central/Northern)
    {
        "code": 13, "name_arabic": "تلمسان", "name_french": "Tlemcen",
        "capital": "Tlemcen", "area_km2": 9061, "population": 945525,
        "districts": 20, "communes": 54,
        "coordinates": {"lat": 34.88, "lng": -1.32},
        "region": "North (Coastal)", "established": "1968",
        "economy": ["Tourism", "Agriculture", "Trade", "Light industry"],
        "landmarks": ["Great Mosque of Tlemcen", "El Ourit waterfall", "Lalla Setti plateau"]
    },
    {
        "code": 14, "name_arabic": "تيارت", "name_french": "Tiaret",
        "capital": "Tiaret", "area_km2": 20673, "population": 842063,
        "districts": 14, "communes": 42,
        "coordinates": {"lat": 35.37, "lng": 1.32},
        "region": "Highlands", "established": "1968",
        "economy": ["Agriculture", "Livestock", "Light industry"],
        "landmarks": ["Edrak ruins", "Djedid fortress", "Agricultural plains"]
    },
    {
        "code": 15, "name_arabic": "تيزي وزو", "name_french": "Tizi Ouzou",
        "capital": "Tizi Ouzou", "area_km2": 2993, "population": 1119646,
        "districts": 21, "communes": 67,
        "coordinates": {"lat": 36.72, "lng": 4.06},
        "region": "North (Kabylie)", "established": "1968",
        "economy": ["Commerce", "Light industry", "Transportation"],
        "landmarks": ["Djurdjura Mountains", "Kabylie region", "Douera valley"]
    },
    {
        "code": 16, "name_arabic": "الجزائر", "name_french": "Algiers",
        "capital": "Algiers", "area_km2": 1190, "population": 7796923,
        "districts": 13, "communes": 57,
        "coordinates": {"lat": 36.75, "lng": 3.05},
        "region": "North (Coastal/Capital)", "established": "1968",
        "economy": ["Government", "Finance", "Port activities", "Services", "Industry"],
        "landmarks": ["Casbah of Algiers (UNESCO)", "Notre Dame d'Afrique", "Kama park"]
    },
    {
        "code": 17, "name_arabic": "الجلفة", "name_french": "Djelfa",
        "capital": "Djelfa", "area_km2": 66415, "population": 1223223,
        "districts": 12, "communes": 36,
        "coordinates": {"lat": 34.67, "lng": 3.38},
        "region": "Highlands (Steppe)", "established": "1974",
        "economy": ["Agriculture (cereal)", "Livestock breeding", "Trade"],
        "landmarks": ["Salt lakes (chotts)", "Hauts Plateaux", "Roman ruins"]
    },
    {
        "code": 18, "name_arabic": "جيجل", "name_french": "Jijel",
        "capital": "Jijel", "area_km2": 2577, "population": 634412,
        "districts": 11, "communes": 28,
        "coordinates": {"lat": 36.80, "lng": 5.76},
        "region": "North (Coastal)", "established": "1974",
        "economy": ["Fishing", "Tourism", "Agriculture", "Food processing"],
        "landmarks": ["Corniche jijelienne", "Caves of Aquafortis", "Beach resorts"]
    },
    {
        "code": 19, "name_arabic": "سطيف", "name_french": "Sétif",
        "capital": "Sétif", "area_km2": 6504, "population": 1496150,
        "districts": 20, "communes": 60,
        "coordinates": {"lat": 36.19, "lng": 5.41},
        "region": "Highlands (East)", "established": "1968",
        "economy": ["Agriculture (cereal)", "Trade", "Industry", "Services"],
        "landmarks": ["Roman circus of Sitifis", "Guellal zoo", "Aurès foothills"]
    },
    {
        "code": 20, "name_arabic": "سعيدة", "name_french": "Saïda",
        "capital": "Saïda", "area_km2": 6764, "population": 328685,
        "districts": 6, "communes": 16,
        "coordinates": {"lat": 34.83, "lng": 0.15},
        "region": "Highlands", "established": "1968",
        "economy": ["Agriculture", "Forestry", "Water resources"],
        "landmarks": ["Djebel Dhaya", "Thermal springs", "Forest reserves"]
    },
    {
        "code": 21, "name_arabic": "سكيكدة", "name_french": "Skikda",
        "capital": "Skikda", "area_km2": 4026, "population": 1109355,
        "districts": 13, "communes": 38,
        "coordinates": {"lat": 36.87, "lng": 6.91},
        "region": "North (Coastal)", "established": "1974",
        "economy": ["Petrochemical industry", "Port activities", "Refining", "Fishing"],
        "landmarks": ["Natural gas liquefaction plant", "Coastal beaches", "Fort de l'Eau"]
    },
    {
        "code": 22, "name_arabic": "سيدي بلعباس", "name_french": "Sidi Bel Abbès",
        "capital": "Sidi Bel Abbès", "area_km2": 9151, "population": 603369,
        "districts": 15, "communes": 52,
        "coordinates": {"lat": 34.85, "lng": -0.60},
        "region": "North (Inland)", "established": "1968",
        "economy": ["Agriculture", "Military base", "Trade", "Services"],
        "landmarks": ["French Foreign Legion historical site", "Agricultural plains"]
    },
    {
        "code": 23, "name_arabic": "عنابة", "name_french": "Annaba",
        "capital": "Annaba", "area_km2": 1439, "population": 640050,
        "districts": 6, "communes": 12,
        "coordinates": {"lat": 36.93, "lng": 7.77},
        "region": "North (Coastal)", "established": "1968",
        "economy": ["Steel industry (ArcelorMittal)", "Port activities", "Mining"],
        "landmarks": ["Ruins of Hippo Regius", "Basilica of Saint Augustine", "Mediterranean coast"]
    },
    {
        "code": 24, "name_arabic": "قالمة", "name_french": "Guelma",
        "capital": "Guelma", "area_km2": 4101, "population": 482261,
        "districts": 10, "communes": 34,
        "coordinates": {"lat": 36.46, "lng": 7.43},
        "region": "North East", "established": "1974",
        "economy": ["Agriculture", "Textile industry", "Iron mining"],
        "landmarks": ["Roman theater of Calama", "Hot springs (hammam)", "Monts des Nemencha"]
    },
    {
        "code": 25, "name_arabic": "قسنطينة", "name_french": "Constantine",
        "capital": "Constantine", "area_km2": 2187, "population": 1012643,
        "districts": 6, "communes": 12,
        "coordinates": {"lat": 36.37, "lng": 6.62},
        "region": "Highlands (East)", "established": "1968",
        "economy": ["Industry", "Trade", "Education (university hub)", "Services"],
        "landmarks": ["Pont Sidi Rached", "Rhumel River gorges", "Palace of Ahmed Bey"]
    },
    {
        "code": 26, "name_arabic": "المدية", "name_french": "Médéa",
        "capital": "Médéa", "area_km2": 8866, "population": 830943,
        "districts": 19, "communes": 64,
        "coordinates": {"lat": 36.27, "lng": 2.75},
        "region": "Highlands (Tell Atlas)", "established": "1968",
        "economy": ["Agriculture", "Breeding", "Crafts", "Trade"],
        "landmarks": ["Tell Atlas Mountains", "Chahrazad plains", "Oued Chelif valley"]
    },
    
    # Wilayas 27-40 (Agent 3: Eastern/Central)
    {
        "code": 27, "name_arabic": "مستغانم", "name_french": "Mostaganem",
        "capital": "Mostaganem", "area_km2": 2269, "population": 746947,
        "districts": 10, "communes": 32,
        "coordinates": {"lat": 35.933, "lng": 0.083},
        "region": "North (Coastal)", "established": "1984",
        "economy": ["Agriculture", "Fishing", "Port activities", "Education"],
        "landmarks": ["University of Mostaganem", "Salah Bey Mosque", "Coastal beaches"]
    },
    {
        "code": 28, "name_arabic": "المسيلة", "name_french": "M'Sila",
        "capital": "M'Sila", "area_km2": 18718, "population": 991846,
        "districts": 15, "communes": 47,
        "coordinates": {"lat": 35.71, "lng": 4.53},
        "region": "Highlands", "established": "1984",
        "economy": ["Agriculture (cereal)", "Pastoralism", "Traditional crafts"],
        "landmarks": ["Maadid ksar", "Jebel Messaad", "Salt flats"]
    },
    {
        "code": 29, "name_arabic": "معسكر", "name_french": "Mascara",
        "capital": "Mascara", "area_km2": 5941, "population": 784073,
        "districts": 16, "communes": 43,
        "coordinates": {"lat": 35.40, "lng": 0.15},
        "region": "North Inland", "established": "1984",
        "economy": ["Agriculture", "Olive cultivation", "Livestock"],
        "landmarks": ["Castle of Mascara", "Tijdit ruins", "Olive groves"]
    },
    {
        "code": 30, "name_arabic": "ورقلة", "name_french": "Ouargla",
        "capital": "Ouargla", "area_km2": 211980, "population": 558558,
        "districts": 10, "communes": 21,
        "coordinates": {"lat": 33.81, "lng": 5.32},
        "region": "Sahara", "established": "1974",
        "economy": ["Oil & gas production", "Date palm cultivation", "Tourism"],
        "landmarks": ["Hassi Messaoud oil field", "Oasis complexes", "Desert landscapes"]
    },
    {
        "code": 31, "name_arabic": "وهران", "name_french": "Oran",
        "capital": "Oran", "area_km2": 2121, "population": 1584607,
        "districts": 9, "communes": 26,
        "coordinates": {"lat": 35.69, "lng": -0.63},
        "region": "North (Coastal)", "established": "1968",
        "economy": ["Industry", "Port activities", "Commerce", "Tourism"],
        "landmarks": ["Santa Cruz chapel", "Port of Oran", "Plage des Aiguilles"]
    },
    {
        "code": 32, "name_arabic": "البيض", "name_french": "El Bayadh",
        "capital": "El Bayadh", "area_km2": 78870, "population": 262187,
        "districts": 8, "communes": 22,
        "coordinates": {"lat": 33.68, "lng": 0.97},
        "region": "Highlands/Sahara fringe", "established": "1984",
        "economy": ["Pastoralism (sheep/goats)", "Agriculture (steppe)", "Traditional crafts"],
        "landmarks": ["Steppe landscape", "Sheep herding areas", "Atlas foothills"]
    },
    {
        "code": 33, "name_arabic": "إيليزي", "name_french": "Illizi",
        "capital": "Illizi", "area_km2": 284618, "population": 54490,
        "districts": 3, "communes": 6,
        "coordinates": {"lat": 32.54, "lng": 8.28},
        "region": "Sahara (Southeast)", "established": "1984",
        "economy": ["Oil & gas exploration", "Desert tourism", "Trans-Saharan trade"],
        "landmarks": ["Tassili n'Ajjer (UNESCO)", "Rock art sites", "Tadrart Rouge"]
    },
    {
        "code": 34, "name_arabic": "برج بوعريريج", "name_french": "Bordj Bou Arréridj",
        "capital": "Bordj Bou Arréridj", "area_km2": 4115, "population": 634396,
        "districts": 10, "communes": 34,
        "coordinates": {"lat": 36.07, "lng": 4.75},
        "region": "North Inland", "established": "1984",
        "economy": ["Agriculture", "Trade", "Textile industry", "Construction materials"],
        "landmarks": ["Borj fortress", "Agricultural plains", "Commercial center"]
    },
    {
        "code": 35, "name_arabic": "بومرداس", "name_french": "Boumerdès",
        "capital": "Boumerdès", "area_km2": 1591, "population": 802083,
        "districts": 9, "communes": 32,
        "coordinates": {"lat": 36.77, "lng": 3.47},
        "region": "North (Coastal)", "established": "1984",
        "economy": ["IT services", "Education (universities)", "Tourism", "Fishing"],
        "landmarks": ["University complex", "Coastal resorts", "Mountains backdrop"]
    },
    {
        "code": 36, "name_arabic": "الطارف", "name_french": "El Tarf",
        "capital": "El Tarf", "area_km2": 3339, "population": 411783,
        "districts": 7, "communes": 24,
        "coordinates": {"lat": 36.75, "lng": 8.13},
        "region": "North East (Coastal)", "established": "1984",
        "economy": ["Fishing", "Agriculture (citrus)", "Tourism", "Border trade"],
        "landmarks": ["El Kala National Park", "Lake Tonga", "Tunisia border crossing"]
    },
    {
        "code": 37, "name_arabic": "تندوف", "name_french": "Tindouf",
        "capital": "Tindouf", "area_km2": 159000, "population": 49149,
        "districts": 1, "communes": 2,
        "coordinates": {"lat": 27.67, "lng": -8.13},
        "region": "Sahara (Southwest)", "established": "1974",
        "economy": ["Phosphate mining", "Humanitarian aid logistics", "Trade"],
        "landmarks": ["Sahrawi refugee camps area", "Hamada desert", "Border regions"]
    },
    {
        "code": 38, "name_arabic": "تيسمسيلت", "name_french": "Tissemsilt",
        "capital": "Tissemsilt", "area_km2": 3152, "population": 296366,
        "districts": 8, "communes": 22,
        "coordinates": {"lat": 35.61, "lng": 1.82},
        "region": "North Inland", "established": "1984",
        "economy": ["Agriculture", "Forestry", "Hydroelectric power"],
        "landmarks": ["Boukornine dam", "Forest areas", "Mountain terrain"]
    },
    {
        "code": 39, "name_arabic": "الوادي", "name_french": "El Oued",
        "capital": "El Oued", "area_km2": 54573, "population": 673934,
        "districts": 12, "communes": 30,
        "coordinates": {"lat": 33.51, "lng": 6.87},
        "region": "Sahara (Northeast)", "established": "1984",
        "economy": ["Date palm cultivation", "Agriculture (oases)", "Trade", "Rug weaving"],
        "landmarks": ["Golden sand dunes", "Oasis towns", "Traditional architecture"]
    },
    {
        "code": 40, "name_arabic": "خنشلة", "name_french": "Khenchela",
        "capital": "Khenchela", "area_km2": 9811, "population": 386683,
        "districts": 8, "communes": 21,
        "coordinates": {"lat": 35.43, "lng": 7.15},
        "region": "Aurès Mountains", "established": "1984",
        "economy": ["Agriculture", "Forestry", "Tourism (thermal springs)"],
        "landmarks": ["Hamam Essalihine (Roman baths)", "Aurès Mountains", "Cedar forests"]
    },
    
    # Wilayas 41-52 (Agent 4: Southern + newer 2019 wilayas)
    {
        "code": 41, "name_arabic": "سوق أهراس", "name_french": "Souk Ahras",
        "capital": "Souk Ahras", "area_km2": 4541, "population": 440299,
        "districts": 10, "communes": 26,
        "coordinates": {"lat": 36.29, "lng": 7.96},
        "region": "North East (Aurès)", "established": "1984",
        "economy": ["Agriculture", "Trade with Tunisia", "Agro-food industry"],
        "landmarks": ["Birthplace of Saint Augustine", "Ancient Thagaste ruins", "Tunisia border"]
    },
    {
        "code": 42, "name_arabic": "تيبازة", "name_french": "Tipaza",
        "capital": "Tipaza", "area_km2": 2166, "population": 617661,
        "districts": 10, "communes": 28,
        "coordinates": {"lat": 36.59, "lng": 2.53},
        "region": "North (Coastal)", "established": "1984",
        "economy": ["Tourism (Roman ruins)", "Agriculture", "Fishing", "Services"],
        "landmarks": ["Tipaza Roman ruins (UNESCO)", "Chréa National Park extension", "Mediterranean coast"]
    },
    {
        "code": 43, "name_arabic": "ميلة", "name_french": "Mila",
        "capital": "Mila", "area_km2": 9375, "population": 768419,
        "districts": 13, "communes": 34,
        "coordinates": {"lat": 36.45, "lng": 6.27},
        "region": "North East", "established": "1984",
        "economy": ["Agriculture", "Grain production", "Light industry"],
        "landmarks": ["Tigrina reservoir", "Historical granaries", "Agricultural plains"]
    },
    {
        "code": 44, "name_arabic": "عين الدفلى", "name_french": "Aïn Defla",
        "capital": "Khemis Miliana", "area_km2": 4897, "population": 771890,
        "districts": 8, "communes": 36,
        "coordinates": {"lat": 36.26, "lng": 2.23},
        "region": "North", "established": "1984",
        "economy": ["Agriculture", "Petrochemical industry", "Food processing"],
        "landmarks": ["Chelif river valley", "Dams/reservoirs", "Industrial zones"]
    },
    {
        "code": 45, "name_arabic": "النعامة", "name_french": "Naâma",
        "capital": "Naâma", "area_km2": 29950, "population": 209470,
        "districts": 7, "communes": 12,
        "coordinates": {"lat": 33.27, "lng": -0.31},
        "region": "Highlands (Western)", "established": "1984",
        "economy": ["Pastoralism", "Agriculture (steppe crops)", "Border trade"],
        "landmarks": ["Ain Sefra (the pearl of Saoura)", "Atlas Mountains", "Thermal springs"]
    },
    {
        "code": 46, "name_arabic": "عين تموشنت", "name_french": "Aïn Témouchent",
        "capital": "Aïn Témouchent", "area_km2": 2377, "population": 432353,
        "districts": 8, "communes": 28,
        "coordinates": {"lat": 35.31, "lng": -1.13},
        "region": "North West", "established": "1984",
        "economy": ["Agriculture (viticulture)", "Fishing", "Food processing"],
        "landmarks": ["Vineyards", "Coastal beaches", "Marine resorts"]
    },
    {
        "code": 47, "name_arabic": "غرداية", "name_french": "Ghardaïa",
        "capital": "Ghardaïa", "area_km2": 21224, "population": 391671,
        "districts": 9, "communes": 13,
        "coordinates": {"lat": 32.49, "lng": 3.67},
        "region": "Sahara (M'Zab Valley)", "established": "1984",
        "economy": ["Date palm cultivation", "Traditional crafts", "Tourism"],
        "landmarks": ["M'Zab Valley (UNESCO)", "Pentapolis cities", "Mozabite culture"]
    },
    {
        "code": 48, "name_arabic": "ريزان", "name_french": "Relizane",
        "capital": "Relizane", "area_km2": 4870, "population": 733060,
        "districts": 7, "communes": 38,
        "coordinates": {"lat": 35.94, "lng": 0.33},
        "region": "North West", "established": "1984",
        "economy": ["Agriculture", "Food industry", "Hydroelectric dams"],
        "landmarks": ["Chelif valley", "Dams (Youssef, Hamiz)", "Agricultural lands"]
    },
    {
        "code": 49, "name_arabic": "المغير", "name_french": "El M'Ghair",
        "capital": "El M'Ghair", "area_km2": 8835, "population": 162267,
        "districts": 6, "communes": 12,
        "coordinates": {"lat": 33.41, "lng": 6.66},
        "region": "Sahara", "established": "2019",
        "economy": ["Date palm cultivation", "Agriculture (oases)", "Pastoralism"],
        "landmarks": ["Oasis complexes", "Salt flats (chotts)", "Desert landscape"]
    },
    {
        "code": 50, "name_arabic": "المنيعة", "name_french": "El Meniaa",
        "capital": "El Meniaa (El Golea)", "area_km2": 62215, "population": 57276,
        "districts": 4, "communes": 8,
        "coordinates": {"lat": 30.57, "lng": 2.87},
        "region": "Sahara", "established": "2019",
        "economy": ["Date palm cultivation", "Tourism (oasis)", "Agriculture"],
        "landmarks": ["El Golea oasis (Pearl of Desert)", "Zousfana wadi", "Desert routes"]
    },
    {
        "code": 51, "name_arabic": "أولاد جلال", "name_french": "Ouled Djellal",
        "capital": "Ouled Djellal", "area_km2": 11410, "population": 174219,
        "districts": 6, "communes": 12,
        "coordinates": {"lat": 35.81, "lng": 5.91},
        "region": "Sahara (Highlands fringe)", "established": "2019",
        "economy": ["Sheep/cattle breeding", "Agriculture", "Trade"],
        "landmarks": ["Hodna mountains", "Pastoral lands", "Trans-Saharan route"]
    },
    {
        "code": 52, "name_arabic": "برج باجي مختار", "name_french": "Bordj Badji Mokhtar",
        "capital": "Bordj Badji Mokhtar", "area_km2": 120026, "population": 16437,
        "districts": 2, "communes": 4,
        "coordinates": {"lat": 21.94, "lng": -0.95},
        "region": "Sahara (Deep South)", "established": "2019",
        "economy": ["Trans-Saharan trade", "Nomadic pastoralism", "Border logistics"],
        "landmarks": ["Mali/Niger borders", "Trans-Saharan road", "Desert wilderness"]
    },
    
    # Wilayas 53-58 (Agent 5: Newest 2019 wilayas)
    {
        "code": 53, "name_arabic": "بني عباس", "name_french": "Béni Abbès",
        "capital": "Béni Abbès", "area_km2": 101350, "population": 50163,
        "districts": 6, "communes": 10,
        "coordinates": {"lat": 30.13, "lng": -2.17},
        "region": "Sahara (Saoura)", "established": "2019",
        "economy": ["Date palm cultivation", "Tourism (oasis)", "Agriculture"],
        "landmarks": ["Beni Abbes oasis", "Pyramid-shaped house", "Saoura region"],
        "parent_wilaya": "Béchar"
    },
    {
        "code": 54, "name_arabic": "تيميمون", "name_french": "Timimoun",
        "capital": "Timimoun", "area_km2": 65203, "population": 122019,
        "districts": 4, "communes": 10,
        "coordinates": {"lat": 29.29, "lng": 0.23},
        "region": "Sahara (Gourara)", "established": "2019",
        "economy": ["Date palm cultivation", "Tourism (red dunes)", "Agriculture"],
        "landmarks": ["Red ochre architecture", "Timimoun palm grove", "Salt lake"],
        "parent_wilaya": "Adrar"
    },
    {
        "code": 55, "name_arabic": "تقرت", "name_french": "Touggourt",
        "capital": "Touggourt", "area_km2": 17428, "population": 247221,
        "districts": 4, "communes": 11,
        "coordinates": {"lat": 33.11, "lng": 6.07},
        "region": "Sahara (Oued Righ)", "established": "2019",
        "economy": ["Oil & gas production", "Date palm cultivation", "Trade"],
        "landmarks": ["Large palmeries", "Oil fields", "Oasis towns"],
        "parent_wilaya": "Ouargla"
    },
    {
        "code": 56, "name_arabic": "جانت", "name_french": "Djanet",
        "capital": "Djanet", "area_km2": 86185, "population": 17618,
        "districts": 2, "communes": 3,
        "coordinates": {"lat": 24.55, "lng": 9.48},
        "region": "Sahara (Tassili)", "established": "2019",
        "economy": ["Desert tourism", "Tuareg culture", "Trans-Saharan trade"],
        "landmarks": ["Tassili n'Ajjer (UNESCO)", "Rock art", "Tuareg cultural center"],
        "parent_wilaya": "Illizi"
    },
    {
        "code": 57, "name_arabic": "عين صالح", "name_french": "In Salah",
        "capital": "In Salah", "area_km2": 131220, "population": 50392,
        "districts": 2, "communes": 3,
        "coordinates": {"lat": 27.25, "lng": 2.47},
        "region": "Sahara (Tidikelt)", "established": "2019",
        "economy": ["Gas industry (Sonatrach)", "Trans-Saharan trade", "Renewable energy"],
        "landmarks": ["In Salah gas facility", "Tidikelt oases", "Desert crossroads"],
        "parent_wilaya": "Tamanrasset"
    },
    {
        "code": 58, "name_arabic": "عين قزّام", "name_french": "In Guezzam",
        "capital": "In Guezzam", "area_km2": 88126, "population": 11202,
        "districts": 2, "communes": 2,
        "coordinates": {"lat": 23.51, "lng": 5.74},
        "region": "Sahara (Far South)", "established": "2019",
        "economy": ["Border trade (Niger)", "Logistics", "Pastoralism"],
        "landmarks": ["Niger border crossing", "Tassili slopes", "Southernmost point DZ"],
        "parent_wilaya": "Tamanrasset"
    }
]


def create_wilayas_workbook():
    """Create comprehensive Excel workbook with all 58 Algerian wilayas data."""
    
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # ===== Sheet 1: Complete Data =====
    ws_data = wb.create_sheet("Wilayas Complete Data", 0)
    setup_worksheet_data(ws_data)
    
    # ===== Sheet 2: Regional Summary =====
    ws_region = wb.create_sheet("Regional Summary")
    setup_worksheet_regional(ws_region)
    
    # ===== Sheet 3: Statistics Dashboard =====
    ws_stats = wb.create_sheet("Statistics")
    setup_worksheet_statistics(ws_stats)
    
    # ===== Sheet 4: Data Quality Notes =====
    ws_notes = wb.create_sheet("Data Notes")
    setup_worksheet_notes(ws_notes)
    
    # Delete default sheet
    wb.remove(default_sheet)
    
    return wb


def setup_worksheet_data(ws):
    """Setup main data sheet with all 58 wilayas."""
    
    # Setup sheet basics
    setup_sheet(ws, title="Complete Database of Algeria's 58 Wilayas (Administrative Provinces)", last_col=14)
    
    # Define headers
    headers = [
        ("Code", 8),
        ("Arabic Name", 16),
        ("French Name", 20),
        ("Capital City", 18),
        ("Area (km²)", 14),
        ("Population", 14),
        ("Districts", 10),
        ("Communes", 10),
        ("Latitude", 12),
        ("Longitude", 12),
        ("Region", 22),
        ("Established", 12),
        ("Main Economy", 35),
        ("Key Landmarks", 40)
    ]
    
    # Write headers at row 4
    header_row = 4
    for col_idx, (header, width) in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Style header row
    style_header_row(ws, header_row, 2, len(headers) + 1)
    
    # Write data rows
    for row_idx, wilaya in enumerate(ALL_WILAYAS):
        current_row = header_row + 1 + row_idx
        
        # Write data
        data_row = [
            wilaya["code"],
            wilaya["name_arabic"],
            wilaya["name_french"],
            wilaya["capital"],
            wilaya["area_km2"],
            wilaya["population"],
            wilaya["districts"],
            wilaya["communes"],
            wilaya["coordinates"]["lat"],
            wilaya["coordinates"]["lng"],
            wilaya["region"],
            wilaya.get("established", ""),
            ", ".join(wilaya.get("economy", [])[:3]),
            ", ".join(wilaya.get("landmarks", [])[:3])
        ]
        
        for col_idx, value in enumerate(data_row, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            
            # Number formatting for specific columns
            if col_idx == 6:  # Area
                cell.number_format = "#,##0"
            elif col_idx == 7:  # Population
                cell.number_format = "#,##0"
            elif col_idx in [10, 11]:  # Coordinates
                cell.number_format = "0.00"
            
            # Alignment
            if isinstance(value, (int, float)):
                cell.alignment = align_number()
            else:
                cell.alignment = align_text()
        
        # Style data row (alternating colors)
        style_data_row(ws, current_row, 2, len(headers) + 1, row_idx)
    
    # Add totals row
    total_row = header_row + len(ALL_WILAYAS) + 1
    ws.cell(row=total_row, column=2, value="TOTAL / AVERAGE")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{total_row-1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{header_row+1}:G{total_row-1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{header_row+1}:H{total_row-1})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I{header_row+1}:I{total_row-1})")
    ws.cell(row=total_row, column=10, value=f"=AVERAGE(J{header_row+1}:J{total_row-1})")
    ws.cell(row=total_row, column=11, value=f"=AVERAGE(K{header_row+1}:K{total_row-1})")
    
    style_total_row(ws, total_row, 2, len(headers) + 1)


def setup_worksheet_regional(ws):
    """Setup regional summary sheet."""
    
    setup_sheet(ws, title="Regional Breakdown of Algerian Wilayas", last_col=8)
    
    # Calculate regional statistics
    regions = {}
    for w in ALL_WILAYAS:
        region = w["region"]
        if region not in regions:
            regions[region] = {"count": 0, "total_pop": 0, "total_area": 0, "wilayas": []}
        regions[region]["count"] += 1
        regions[region]["total_pop"] += w["population"]
        regions[region]["total_area"] += w["area_km2"]
        regions[region]["wilayas"].append(w["name_french"])
    
    # Headers
    headers = [
        ("Region", 28),
        ("Number of Wilayas", 18),
        ("Total Population", 18),
        ("Total Area (km²)", 18),
        ("Avg Population", 16),
        ("Avg Area (km²)", 16),
        ("% of Total Pop", 14),
        ("Sample Wilayas", 45)
    ]
    
    header_row = 4
    for col_idx, (header, width) in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    style_header_row(ws, header_row, 2, len(headers) + 1)
    
    # Calculate grand totals
    grand_pop = sum(r["total_pop"] for r in regions.values())
    grand_area = sum(r["total_area"] for r in regions.values())
    
    # Sort by population descending
    sorted_regions = sorted(regions.items(), key=lambda x: x[1]["total_pop"], reverse=True)
    
    for row_idx, (region_name, data) in enumerate(sorted_regions):
        current_row = header_row + 1 + row_idx
        
        row_data = [
            region_name,
            data["count"],
            data["total_pop"],
            data["total_area"],
            int(data["total_pop"] / data["count"]),
            int(data["total_area"] / data["count"]),
            round(data["total_pop"] / grand_pop * 100, 1),
            ", ".join(data["wilayas"][:4]) + ("..." if len(data["wilayas"]) > 4 else "")
        ]
        
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if isinstance(value, (int, float)) and col_idx not in [8]:
                cell.alignment = align_number()
                if col_idx in [4, 5, 6, 7]:
                    cell.number_format = "#,##0"
                elif col_idx == 9:
                    cell.number_format = "0.0%"
            else:
                cell.alignment = align_text()
        
        style_data_row(ws, current_row, 2, len(headers) + 1, row_idx)
    
    # Add bar chart for population by region
    chart = create_bar_chart(style=10, width=18, height=10)
    data_ref = Reference(ws, min_col=4, min_row=header_row, max_row=header_row + len(sorted_regions))
    cats_ref = Reference(ws, min_col=2, min_row=header_row + 1, max_row=header_row + len(sorted_regions))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.y_axis.title = 'Population'
    chart.x_axis.title = 'Region'
    setup_chart_titles(chart, "Population Distribution by Region", y_title="Population", x_title="Region")
    apply_chart_colors(chart)
    ws.add_chart(chart, "J4")


def setup_worksheet_statistics(ws):
    """Setup statistics dashboard sheet."""
    
    setup_sheet(ws, title="Key Statistics - Algeria's 58 Wilayas", last_col=6)
    
    # Calculate key statistics
    total_population = sum(w["population"] for w in ALL_WILAYAS)
    total_area = sum(w["area_km2"] for w in ALL_WILAYAS)
    total_districts = sum(w["districts"] for w in ALL_WILAYAS)
    total_communes = sum(w["communes"] for w in ALL_WILAYAS)
    
    avg_population = total_population // len(ALL_WILAYAS)
    avg_area = total_area // len(ALL_WILAYAS)
    
    most_populous = max(ALL_WILAYAS, key=lambda x: x["population"])
    least_populous = min(ALL_WILAYAS, key=lambda x: x["population"])
    largest = max(ALL_WILAYAS, key=lambda x: x["area_km2"])
    smallest = min(ALL_WILAYAS, key=lambda x: x["area_km2"])
    
    # Count by establishment year
    est_years = {}
    for w in ALL_WILAYAS:
        year = w.get("established", "Unknown")
        est_years[year] = est_years.get(year, 0) + 1
    
    # KPI Section Header
    kpi_header_row = 4
    ws.merge_cells(start_row=kpi_header_row, start_column=2, end_row=kpi_header_row, end_column=6)
    kpi_cell = ws.cell(row=kpi_header_row, column=2, value="KEY PERFORMANCE INDICATORS")
    kpi_cell.font = font_subheader()
    kpi_cell.alignment = align_header()
    
    # KPIs
    kpis = [
        ("Total Wilayas", f"{len(ALL_WILAYAS)}"),
        ("Total Population", f"{total_population:,}"),
        ("Total Area", f"{total_area:,} km²"),
        ("Total Districts", f"{total_districts:,}"),
        ("Total Communes", f"{total_communes:,}"),
        ("Average Population/Wilaya", f"{avg_population:,}"),
        ("Average Area/Wilaya", f"{avg_area:,} km²"),
        ("Population Density (avg)", f"{total_population / total_area:.2f} people/km²")
    ]
    
    for idx, (label, value) in enumerate(kpis):
        row = kpi_header_row + 1 + idx
        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.font = font_body()
        label_cell.alignment = align_text()
        
        value_cell = ws.cell(row=row, column=4, value=value)
        value_cell.font = font_subheader()
        value_cell.alignment = align_number()
    
    # Extremes Section
    extremes_header = kpi_header_row + len(kpis) + 2
    ws.merge_cells(start_row=extremes_header, start_column=2, end_row=extremes_header, end_column=6)
    ext_cell = ws.cell(row=extremes_header, column=2, value="EXTREME VALUES")
    ext_cell.font = font_subheader()
    ext_cell.alignment = align_header()
    
    extremes = [
        ("Most Populous", f"{most_populous['name_french']} ({most_populous['population']:,})"),
        ("Least Populous", f"{least_populous['name_french']} ({least_populous['population']:,})"),
        ("Largest Area", f"{largest['name_french']} ({largest['area_km2']:,} km²)"),
        ("Smallest Area", f"{smallest['name_french']} ({smallest['area_km2']:,} km²)")
    ]
    
    for idx, (label, value) in enumerate(extremes):
        row = extremes_header + 1 + idx
        ws.cell(row=row, column=2, value=label).font = font_body()
        ws.cell(row=row, column=4, value=value).font = font_body()
    
    # Establishment Timeline
    timeline_header = extremes_header + len(extremes) + 2
    ws.merge_cells(start_row=timeline_header, start_column=2, end_row=timeline_header, end_column=6)
    tl_cell = ws.cell(row=timeline_header, column=2, value="ESTABLISHMENT TIMELINE")
    tl_cell.font = font_subheader()
    tl_cell.alignment = align_header()
    
    # Headers for timeline table
    th_row = timeline_header + 1
    ws.cell(row=th_row, column=2, value="Year Established").font = font_header()
    ws.cell(row=th_row, column=3, value="Count").font = font_header()
    ws.cell(row=th_row, column=4, value="% of Total").font = font_header()
    
    sorted_years = sorted(est_years.items(), key=lambda x: x[0])
    for idx, (year, count) in enumerate(sorted_years):
        row = th_row + 1 + idx
        ws.cell(row=row, column=2, value=year).font = font_body()
        ws.cell(row=row, column=3, value=count).font = font_body()
        ws.cell(row=row, column=4, value=f"{count/len(ALL_WILAYAS)*100:.1f}%").font = font_body()
    
    # Column widths
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 15


def setup_worksheet_notes(ws):
    """Setup data quality notes sheet."""
    
    setup_sheet(ws, title="Data Quality Notes & Sources", last_col=4)
    
    notes_content = [
        "",
        "DATA SOURCES",
        "- Wikipedia (English/French/Arabic) - Primary source for administrative data",
        "- Official Algerian government sources (ONS, Ministry of Interior)",
        "- statoids.com - Cross-referenced population and area figures",
        "- Official MTA (Ministère du Tourisme et de l'Artisanat) websites",
        "- CityPopulation.de - Census data verification",
        "",
        "DATA QUALITY NOTES",
        "",
        "Population Data:",
        "- Figures primarily from 2008 census (latest consistently available)",
        "- Some newer wilayas (49-58) have estimated/provisional figures",
        "- Current populations likely differ due to migration and development",
        "",
        "Area Measurements:",
        "- From official administrative records via Wikipedia",
        "- Minor discrepancies exist between sources for some wilayas",
        "- Particularly for large Saharan wilayas where boundaries are complex",
        "",
        "Administrative Evolution:",
        "- Original 15 departments (French colonial era)",
        "- Expanded to 31 wilayas (1974)",
        "- Further expanded to 48 wilayas (1984)",
        "- Latest expansion to 58 wilayas (2019)",
        "- Wilayas 49-58 were created from existing provinces (see parent_wilaya field)",
        "",
        "Coordinate Data:",
        "- Latitude/Longitude represent capital city centers",
        "- Decimal degrees format (WGS84 datum)",
        "",
        "LIMITATIONS",
        "- Some economic activity descriptions may be generalized",
        "- Landmark lists are not exhaustive (top highlights only)",
        "- Newer wilayas (2019) have less historical data available",
        "- Some district/commune counts may reflect recent reorganizations",
        "",
        "GENERATION INFO",
        f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "- Method: Parallel web scraping using 5 concurrent agents",
        "- Total records: 58 wilayas (complete coverage)",
        "- Fields per record: 14 data points"
    ]
    
    for idx, line in enumerate(notes_content):
        row = 4 + idx
        cell = ws.cell(row=row, column=2, value=line)
        
        if line.endswith(":") and line != "":
            cell.font = font_subheader()
        elif line.startswith("-"):
            cell.font = font_body()
        else:
            cell.font = font_body() if line else None
        
        cell.alignment = align_text()
    
    # Merge for better display
    ws.merge_cells(start_row=4, start_column=2, end_row=len(notes_content)+3, end_column=6)
    ws.column_dimensions["B"].width = 100


def main():
    """Main function to generate the Excel file."""
    
    print("=" * 60)
    print("ALGERIA WILAYAS DATA COMPILER")
    print("=" * 60)
    print(f"Compiling data for {len(ALL_WILAYAS)} wilayas...")
    
    # Create workbook
    wb = create_wilayas_workbook()
    
    # Ensure output directory exists
    output_dir = "/home/z/my-project/download"
    os.makedirs(output_dir, exist_ok=True)
    
    # Save file
    output_path = os.path.join(output_dir, "algeria_58_wilayas_complete_dataset.xlsx")
    wb.save(output_path)
    
    print(f"\n✅ Excel file saved successfully!")
    print(f"📁 Location: {output_path}")
    print(f"\n📊 Sheets included:")
    print(f"   1. Wilayas Complete Data - All 58 provinces with 14 data fields each")
    print(f"   2. Regional Summary - Grouped analysis by geographic region")
    print(f"   3. Statistics - KPIs, extreme values, establishment timeline")
    print(f"   4. Data Notes - Sources, quality info, limitations")
    
    return output_path


if __name__ == "__main__":
    main()
